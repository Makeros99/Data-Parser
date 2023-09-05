from openpyxl import Workbook
from openpyxl.worksheet.filters import AutoFilter
import tkinter as tk
from tkinter import filedialog
import time

print("Lütfen Dosya Seçiniz...")
time.sleep(1)
# Tkinter penceresini oluştur
root = tk.Tk()
root.withdraw()  # Pencere görünmez hale getirilir

# Gözat penceresini aç ve dosya seçimini al
filePath = filedialog.askopenfilename()

# Dosya seçilmediyse çık
if not filePath:
    print("Dosya seçilmedi.")
else:
    print(f"Seçilen dosya yolu: {filePath}")

root.destroy()

#Gruplandırmak istediğiniz anahtar kelimeyi aşağıya giriniz.
firstDataString = "rainy"
secondDataString = "foggy"


# Dosyanın içeriğini okuma
with open(filePath, "r") as dosya:
    allIndex = dosya.readlines()

# girilen kelimeye göre verileri sınıflandırma
firstData = []
secondData = []
otherData = []

for index in allIndex:
    if secondDataString in index:
        secondData.append(index)
   
    elif firstDataString in index:
        firstData.append(index)
   
    else:
        otherData.append(index)
        

# Başlıkları belirleme   -------------   Başlıkları Projenize Göre Özelleştiriniz.
firstDataTitle = ["condition","veri1","veri2","veri3","veri4"]
secondDataTitle = ["condition","bilgi1","bilgi2","bilgi3","bilgi4"]
otherDataTitle = ["information1","information2","information3","information4"]
# Excel dosyası oluşturma (kullanıcıdan alınan adı kullanarak)
excelFileName = f"{filePath[:-4]}.xlsx"  
wb = Workbook()

# Sayfaları oluşturma
firstDataSheet = wb.active
firstDataSheet.title = "first_data"
firstDataSheet.append(firstDataTitle)

secondDataSheet = wb.create_sheet(title="second_data")
secondDataSheet.append(secondDataTitle)

otherDataSheet = wb.create_sheet(title="other_data")  
otherDataSheet.append(otherDataTitle)

# Verileri sayfalara ekleme
for index in firstData:
    indexElements = index.strip().split(",")
    firstDataSheet.append(indexElements)

for index in secondData:
    indexElements = index.strip().split(",")
    secondDataSheet.append(indexElements)

for index in otherData:
    indexElements = index.strip().split(",")
    otherDataSheet.append(indexElements)

# Freeze Panes ayarı: 1. satırı sabitleme
firstDataSheet.freeze_panes = "A2"
secondDataSheet.freeze_panes = "A2"


# Filtre eklemek için AutoFilter özelliğini kullanma
firstDataSheet.auto_filter.ref = firstDataSheet.dimensions
secondDataSheet.auto_filter.ref = secondDataSheet.dimensions

# Sütun genişliklerini ayarlama
for page in wb.sheetnames:
    worksheet = wb[page]
    for column in worksheet.columns:
        max_len = max(len(str(cell.value)) for cell in column)
        adjusted_width = (max_len + 4)
        worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

wb.save(excelFileName)

print(f"'{excelFileName}' adlı Excel dosyası oluşturuldu.")